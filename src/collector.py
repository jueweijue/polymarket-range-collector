#!/usr/bin/env python3
import argparse
import csv
import io
import json
import time
import urllib.request
import urllib.error
import zipfile
from dataclasses import dataclass, asdict
from datetime import datetime, timezone, timedelta
from pathlib import Path
from typing import Dict, List, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

GAMMA_API = "https://gamma-api.polymarket.com"
DATA_API = "https://data-api.polymarket.com"
BINANCE_BASE = "https://data.binance.vision/data/spot/daily/aggTrades/BTCUSDT"
USER_AGENT = "Mozilla/5.0"
UTC = timezone.utc
BJT = timezone(timedelta(hours=8))

HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
YES_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
NO_FILL = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


@dataclass
class MarketInfo:
    slug: str
    start_ts: int
    end_ts: int
    condition_id: str
    yes_token_id: str
    no_token_id: str
    title: str


def log(msg: str):
    print(msg, flush=True)


def fetch_json(url: str):
    req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
    with urllib.request.urlopen(req, timeout=30) as resp:
        return json.loads(resp.read())


def ensure_dir(path: Path):
    path.mkdir(parents=True, exist_ok=True)


def load_config(path: Path):
    return json.loads(path.read_text())


def bjt_to_utc_ts(s: str) -> int:
    dt = datetime.fromisoformat(s)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=BJT)
    else:
        dt = dt.astimezone(BJT)
    return int(dt.astimezone(UTC).timestamp())


def ts_to_iso(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=UTC).isoformat()


def ts_to_api_time(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=UTC).strftime('%Y-%m-%dT%H:%M:%SZ')


def ts_to_bjt_text(ts: int) -> str:
    return datetime.fromtimestamp(ts, tz=BJT).strftime('%Y-%m-%d %H:%M:%S')


def daterange_utc(start_ts: int, end_ts: int):
    current = datetime.fromtimestamp(start_ts, tz=UTC).date()
    end_date = datetime.fromtimestamp(end_ts, tz=UTC).date()
    while current <= end_date:
        yield current
        current += timedelta(days=1)


def market_slug_for_ts(start_ts: int) -> str:
    return f"btc-updown-5m-{start_ts}"


def resolve_markets(start_ts: int, end_ts: int) -> List[MarketInfo]:
    """Resolve BTC Up/Down 5m markets in bulk.

    Old behavior queried one slug per 5-minute slot, which becomes painfully slow
    for month-scale ranges. New behavior pages Gamma `/markets` by day-range and
    filters locally for `btc-updown-5m-`.
    """
    all_markets: Dict[str, MarketInfo] = {}
    day_start_ts = start_ts
    one_day = 86400
    day_index = 0
    total_days = max(1, (end_ts - start_ts + one_day - 1) // one_day)

    while day_start_ts < end_ts:
        day_end_ts = min(day_start_ts + one_day, end_ts)
        offset = 0
        limit = 500
        matched_this_day = 0
        log(f"[prepare] resolving day {day_index + 1}/{total_days}: {ts_to_iso(day_start_ts)} -> {ts_to_iso(day_end_ts)}")

        while True:
            url = (
                f"{GAMMA_API}/markets?limit={limit}&offset={offset}"
                f"&start_date_min={ts_to_api_time(day_start_ts)}&start_date_max={ts_to_api_time(day_end_ts)}"
            )
            data = fetch_json(url)
            if not data:
                break

            for market in data:
                slug = market.get("slug", "")
                if not slug.startswith("btc-updown-5m-"):
                    continue
                try:
                    start_raw = market.get("startDate") or market.get("start_date")
                    market_start_ts = int(datetime.fromisoformat(start_raw.replace("Z", "+00:00")).timestamp())
                    if not (start_ts <= market_start_ts < end_ts):
                        continue
                    token_ids = json.loads(market["clobTokenIds"])
                    info = MarketInfo(
                        slug=slug,
                        start_ts=market_start_ts,
                        end_ts=market_start_ts + 300,
                        condition_id=market["conditionId"],
                        yes_token_id=token_ids[0],
                        no_token_id=token_ids[1],
                        title=market.get("question") or market.get("title") or slug,
                    )
                    if slug not in all_markets:
                        all_markets[slug] = info
                        matched_this_day += 1
                except Exception:
                    continue

            log(f"[prepare]   offset={offset} fetched={len(data)} matched_total={matched_this_day}")
            if len(data) < limit:
                break
            offset += limit
            time.sleep(0.1)

        day_start_ts = day_end_ts
        day_index += 1

    markets = sorted(all_markets.values(), key=lambda m: m.start_ts)
    log(f"[prepare] resolved total btc-updown markets: {len(markets)}")
    return markets


def save_markets(project_dir: Path, markets: List[MarketInfo]):
    out = project_dir / "data" / "markets.json"
    ensure_dir(out.parent)
    out.write_text(json.dumps([asdict(m) for m in markets], ensure_ascii=False, indent=2))


def save_prepare_meta(project_dir: Path, meta: dict):
    path = project_dir / "data" / "prepare_meta.json"
    ensure_dir(path.parent)
    path.write_text(json.dumps(meta, ensure_ascii=False, indent=2))


def load_prepare_meta(project_dir: Path):
    path = project_dir / "data" / "prepare_meta.json"
    if not path.exists():
        return None
    return json.loads(path.read_text())


def load_markets(project_dir: Path) -> List[MarketInfo]:
    path = project_dir / "data" / "markets.json"
    if not path.exists():
        raise FileNotFoundError(f"Missing {path}. Run prepare first.")
    return [MarketInfo(**m) for m in json.loads(path.read_text())]


def binance_zip_name(day) -> str:
    return f"BTCUSDT-aggTrades-{day.isoformat()}.zip"


def download_binance_data(project_dir: Path, start_ts: int, end_ts: int):
    target_dir = project_dir / "data" / "binance"
    ensure_dir(target_dir)
    missing = []
    for day in daterange_utc(start_ts, end_ts):
        filename = binance_zip_name(day)
        target = target_dir / filename
        if target.exists() and target.stat().st_size > 0:
            log(f"[binance] exists: {filename}")
            continue
        url = f"{BINANCE_BASE}/{filename}"
        log(f"[binance] download: {url}")
        req = urllib.request.Request(url, headers={"User-Agent": USER_AGENT})
        try:
            with urllib.request.urlopen(req, timeout=120) as resp:
                target.write_bytes(resp.read())
        except urllib.error.HTTPError as e:
            if e.code == 404:
                log(f"[binance] pending: {filename} not published yet, skip for now")
                missing.append(filename)
                continue
            raise
    return missing


def parse_binance_prices(project_dir: Path, start_ts: int, end_ts: int) -> Dict[int, float]:
    prices = {}
    last_price = None
    binance_dir = project_dir / "data" / "binance"
    for day in daterange_utc(start_ts, end_ts):
        zip_path = binance_dir / binance_zip_name(day)
        if not zip_path.exists():
            raise FileNotFoundError(f"Missing Binance zip: {zip_path}")
        with zipfile.ZipFile(zip_path) as zf:
            inner_name = zf.namelist()[0]
            with zf.open(inner_name) as f:
                wrapper = io.TextIOWrapper(f, encoding="utf-8")
                reader = csv.reader(wrapper)
                for row in reader:
                    if not row:
                        continue
                    price = float(row[1])
                    # Binance aggTrades daily zip currently uses event time in microseconds.
                    ts_us = int(row[5])
                    ts = ts_us // 1_000_000
                    if ts < start_ts - 5 or ts > end_ts + 5:
                        continue
                    prices[ts] = price
    result = {}
    for ts in range(start_ts, end_ts + 1):
        if ts in prices:
            last_price = prices[ts]
        result[ts] = last_price
    return result


def trade_key(trade):
    return (
        trade.get("transactionHash", ""),
        str(trade.get("timestamp", "")),
        str(trade.get("price", "")),
        str(trade.get("size", "")),
        trade.get("side", ""),
        trade.get("outcome", ""),
        trade.get("proxyWallet", ""),
    )


def dedupe_trades(trades: List[dict]) -> List[dict]:
    seen = set()
    unique = []
    for trade in trades:
        key = trade_key(trade)
        if key in seen:
            continue
        seen.add(key)
        unique.append(trade)
    return unique


def fetch_trades_page(condition_id: str, side: str = None, filter_amount: float = 0, offset: int = 0, limit: int = 1000):
    url = f"{DATA_API}/trades?market={condition_id}&limit={limit}&offset={offset}"
    if side:
        url += f"&side={side}"
    if filter_amount > 0:
        url += f"&filterType=CASH&filterAmount={filter_amount}"
    return fetch_json(url)


def fetch_trades_for_side(condition_id: str, side: str, filter_amount: float = 0, limit: int = 1000, max_offset: int = 10000) -> Tuple[List[dict], bool]:
    trades = []
    offset = 0
    truncated = False
    while offset <= max_offset:
        try:
            data = fetch_trades_page(condition_id, side=side, filter_amount=filter_amount, offset=offset, limit=limit)
        except Exception as e:
            log(f"[trades] warning side={side} filter={filter_amount} offset={offset}: {e}")
            truncated = True
            break
        if not data:
            break
        trades.extend(data)
        if len(data) < limit:
            break
        next_offset = offset + limit
        if next_offset > max_offset:
            truncated = True
            break
        offset = next_offset
        time.sleep(0.15)
    return trades, truncated


def fetch_market_trades(condition_id: str, filter_amount: float = 0):
    recovery_filters = [] if filter_amount > 0 else [1, 2, 5, 10, 20, 50, 100, 200, 500, 1000]
    all_trades = []
    unresolved = []
    diagnostics = {"base_filter": filter_amount, "sides": {}, "unresolved_sides": unresolved}
    for side in ("BUY", "SELL"):
        side_diag = {"base_truncated": False, "base_count": 0, "recovery_attempts": [], "resolved": True}
        side_trades, truncated = fetch_trades_for_side(condition_id, side=side, filter_amount=filter_amount)
        side_diag["base_truncated"] = truncated
        side_diag["base_count"] = len(side_trades)
        recovered = list(side_trades)
        resolved_by_recovery = not truncated
        if truncated and recovery_filters:
            for recovery_filter in recovery_filters:
                extra_trades, extra_truncated = fetch_trades_for_side(condition_id, side=side, filter_amount=recovery_filter)
                side_diag["recovery_attempts"].append({"filter": recovery_filter, "count": len(extra_trades), "truncated": extra_truncated})
                recovered.extend(extra_trades)
                if not extra_truncated:
                    resolved_by_recovery = True
        recovered = dedupe_trades(recovered)
        diagnostics["sides"][side] = side_diag
        side_diag["resolved"] = resolved_by_recovery
        if truncated and not resolved_by_recovery:
            unresolved.append(side)
        all_trades.extend(recovered)
    unique = dedupe_trades(all_trades)
    diagnostics["unique_total"] = len(unique)
    diagnostics["complete"] = len(unresolved) == 0
    return unique, diagnostics


def trade_cache_path(project_dir: Path, slug: str) -> Path:
    return project_dir / "data" / "trades" / f"{slug}.json"


def save_market_trades(project_dir: Path, market: MarketInfo, trades: List[dict], diagnostics: dict):
    out_dir = project_dir / "data" / "trades"
    ensure_dir(out_dir)
    payload = {
        "market": asdict(market),
        "diagnostics": diagnostics,
        "trades": sorted(trades, key=lambda t: int(t["timestamp"]))
    }
    out_path = trade_cache_path(project_dir, market.slug)
    out_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2))


def load_market_trades(project_dir: Path, slug: str):
    path = trade_cache_path(project_dir, slug)
    return json.loads(path.read_text())


def build_trade_timeline(trades: List[dict], start_ts: int, end_ts: int):
    by_second = {}
    last_yes = None
    last_no = None
    last_yes_ts = None
    last_no_ts = None
    sorted_trades = sorted(trades, key=lambda t: (int(t["timestamp"]), t.get("transactionHash", "")))
    idx = 0
    n = len(sorted_trades)
    for current_ts in range(start_ts, end_ts + 1):
        while idx < n and int(sorted_trades[idx]["timestamp"]) <= current_ts:
            trade = sorted_trades[idx]
            outcome = trade.get("outcome")
            price = float(trade.get("price", 0))
            ts = int(trade["timestamp"])
            if outcome == "Up":
                last_yes = price
                last_yes_ts = ts
            elif outcome == "Down":
                last_no = price
                last_no_ts = ts
            idx += 1
        by_second[current_ts] = {
            "yes_price": last_yes,
            "no_price": last_no,
            "yes_last_trade_ts": last_yes_ts,
            "no_last_trade_ts": last_no_ts,
        }
    return by_second


def autosize(ws):
    for col in ws.columns:
        length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                length = max(length, len(str(cell.value)) if cell.value is not None else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(length + 2, 10), 28)


def write_market_excel(project_dir: Path, market: MarketInfo, diagnostics: dict, btc_prices: Dict[int, float], timeline: Dict[int, dict]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Market Timeline"

    ws.merge_cells('A1:J1')
    ws['A1'] = f"{market.title} ({market.slug})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A2:J2')
    ws['A2'] = f"BJT Window: {ts_to_bjt_text(market.start_ts)} -> {ts_to_bjt_text(market.end_ts)}"

    headers = [
        "Sec", "Time (UTC)", "Time (BJT)", "BTC Price", "YES Price", "NO Price",
        "YES Last Trade (UTC)", "NO Last Trade (UTC)", "YES Age", "NO Age"
    ]
    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGN
        cell.border = THIN_BORDER

    for row_idx, ts in enumerate(range(market.start_ts, market.end_ts + 1), start=5):
        state = timeline[ts]
        yes_ts = state.get('yes_last_trade_ts')
        no_ts = state.get('no_last_trade_ts')
        values = [
            ts - market.start_ts,
            datetime.fromtimestamp(ts, tz=UTC).strftime('%Y-%m-%d %H:%M:%S'),
            ts_to_bjt_text(ts),
            btc_prices.get(ts),
            state.get('yes_price'),
            state.get('no_price'),
            datetime.fromtimestamp(yes_ts, tz=UTC).strftime('%Y-%m-%d %H:%M:%S') if yes_ts else None,
            datetime.fromtimestamp(no_ts, tz=UTC).strftime('%Y-%m-%d %H:%M:%S') if no_ts else None,
            f"{ts - yes_ts}s" if yes_ts else None,
            f"{ts - no_ts}s" if no_ts else None,
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if col_idx == 5 and value is not None:
                cell.fill = YES_FILL
            if col_idx == 6 and value is not None:
                cell.fill = NO_FILL

    autosize(ws)

    summary = wb.create_sheet("Summary")
    summary["A1"] = "Field"
    summary["B1"] = "Value"
    summary["A1"].font = summary["B1"].font = Font(bold=True)
    completeness = "COMPLETE" if diagnostics.get("complete") else f"BEST_EFFORT ({', '.join(diagnostics.get('unresolved_sides', []))})"
    stats = [
        ("Title", market.title),
        ("Slug", market.slug),
        ("Condition ID", market.condition_id),
        ("Start BJT", ts_to_bjt_text(market.start_ts)),
        ("End BJT", ts_to_bjt_text(market.end_ts)),
        ("Total Trades", diagnostics.get("unique_total")),
        ("Completeness", completeness),
        ("YES Base Count", diagnostics.get("sides", {}).get("BUY", {}).get("base_count")),
        ("SELL Base Count", diagnostics.get("sides", {}).get("SELL", {}).get("base_count")),
        ("Data Source", "Binance aggTrades + Polymarket Data API trades"),
        ("YES/NO Method", "Latest observed trade price up to each second"),
    ]
    for idx, (k, v) in enumerate(stats, start=2):
        summary[f"A{idx}"] = k
        summary[f"B{idx}"] = v
        if k == "Completeness" and not diagnostics.get("complete"):
            summary[f"B{idx}"].fill = WARN_FILL
    autosize(summary)

    out_dir = project_dir / "output"
    ensure_dir(out_dir)
    out_path = out_dir / f"{market.slug}.xlsx"
    wb.save(out_path)
    return out_path


def prepare(project_dir: Path, config: dict):
    start_ts = bjt_to_utc_ts(config['range']['start_bjt'])
    end_ts = bjt_to_utc_ts(config['range']['end_bjt'])

    cached_meta = load_prepare_meta(project_dir)
    reuse_prepare = False
    if cached_meta:
        same_range = cached_meta.get('start_ts') == start_ts and cached_meta.get('end_ts') == end_ts
        markets_path = project_dir / 'data' / 'markets.json'
        if same_range and markets_path.exists():
            reuse_prepare = True
            log(f"[prepare] reusing cached market list for range {config['range']['start_bjt']} -> {config['range']['end_bjt']}")

    if reuse_prepare:
        markets = load_markets(project_dir)
    else:
        markets = resolve_markets(start_ts, end_ts)
        if not markets:
            raise RuntimeError('No BTC Up/Down markets resolved for configured range')
        save_markets(project_dir, markets)
        save_prepare_meta(project_dir, {
            'start_ts': start_ts,
            'end_ts': end_ts,
            'start_bjt': config['range']['start_bjt'],
            'end_bjt': config['range']['end_bjt'],
            'market_count': len(markets),
            'prepared_at': datetime.now(tz=UTC).isoformat(),
        })
        log(f"[prepare] markets={len(markets)} saved to data/markets.json")

    missing = download_binance_data(project_dir, start_ts, end_ts)
    if missing:
        log(f"[prepare] binance pending files: {', '.join(missing)}")


def collect_historical(project_dir: Path, config: dict):
    markets = load_markets(project_dir)
    total = len(markets)
    for idx, market in enumerate(markets, start=1):
        cache_path = trade_cache_path(project_dir, market.slug)
        if cache_path.exists() and cache_path.stat().st_size > 0:
            log(f"[history] [{idx}/{total}] skip cached {market.slug}")
            continue
        log(f"[history] [{idx}/{total}] fetch trades for {market.slug}")
        trades, diagnostics = fetch_market_trades(market.condition_id, filter_amount=float(config.get('polymarket', {}).get('filter_amount', 0)))
        save_market_trades(project_dir, market, trades, diagnostics)
        log(f"[history] [{idx}/{total}] {market.slug}: trades={len(trades)} completeness={'ok' if diagnostics.get('complete') else 'best-effort'}")


def export_excels(project_dir: Path, config: dict):
    markets = load_markets(project_dir)
    overall_start = min(m.start_ts for m in markets)
    overall_end = max(m.end_ts for m in markets)
    btc_prices = parse_binance_prices(project_dir, overall_start, overall_end)
    for market in markets:
        payload = load_market_trades(project_dir, market.slug)
        diagnostics = payload['diagnostics']
        trades = payload['trades']
        timeline = build_trade_timeline(trades, market.start_ts, market.end_ts)
        out = write_market_excel(project_dir, market, diagnostics, btc_prices, timeline)
        log(f"[export] wrote {out}")


def main():
    parser = argparse.ArgumentParser(description='Polymarket historical range collector')
    parser.add_argument('--config', required=True, help='Path to config json')
    parser.add_argument('action', choices=['prepare', 'history', 'export', 'all'])
    args = parser.parse_args()

    config_path = Path(args.config).resolve()
    project_dir = config_path.parent.parent.resolve()
    config = load_config(config_path)

    if args.action in ('prepare', 'all'):
        prepare(project_dir, config)
    if args.action in ('history', 'all'):
        collect_historical(project_dir, config)
    if args.action in ('export', 'all'):
        export_excels(project_dir, config)


if __name__ == '__main__':
    main()
